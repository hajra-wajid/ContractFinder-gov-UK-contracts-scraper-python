import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, InvalidSessionIdException
from openpyxl import Workbook, load_workbook

# ===================== CONFIGURATION =====================
FIREFOX_PROFILE_PATH = r"C:\Users\alexuser\AppData\Roaming\Mozilla\Firefox\Profiles\default-release"
FIREFOX_BINARY_PATH = r"C:\Program Files\Mozilla Firefox\firefox.exe"
WEBDRIVER_PATH = r"C:\Users\alexuser\Downloads\alex\geckodriver.exe"
INPUT_EXCEL = "contract_links.xlsx"
OUTPUT_EXCEL = "contract_details.xlsx"
MAX_RETRIES = 7
HEADLESS_MODE = True  # Set to True for headless mode

# Ensure output file exists
if not os.path.exists(OUTPUT_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Title", "Buyer", "Industry", "Location of contract", "Value of contract",
        "Procurement reference", "Published date", "Closing date", "Closing time",
        "Contract start date", "Contract end date", "Contract type", "Procedure type",
        "Contract is suitable for SMEs?", "Contract is suitable for VCSEs?", "Description",
        "Awarded date", "Buyer Contact name", "Buyer Address", "Buyer Email",
        "Website", "Supplier", "Supplier Address", "Reference", "Supplier is SME?", "Supplier is VCSE?"
    ])
    wb.save(OUTPUT_EXCEL)

# ===================== FIREFOX SETUP =====================
options = webdriver.FirefoxOptions()
options.set_preference("profile", FIREFOX_PROFILE_PATH)
options.set_preference("permissions.default.image", 2)  # Disable image loading for faster scraping
options.binary_location = FIREFOX_BINARY_PATH
if HEADLESS_MODE:
    options.add_argument("--headless")

# Initialize WebDriver
service = Service(WEBDRIVER_PATH)
driver = webdriver.Firefox(service=service, options=options)

# ===================== HELPER FUNCTIONS =====================

def get_element_text(selector):
    """Safely get text from an element using CSS selector."""
    try:
        return driver.find_element(By.CSS_SELECTOR, selector).text.strip()
    except NoSuchElementException:
        return ""

def find_value_by_label(label):
    """Finds a value based on its label (e.g., 'Contract end date')."""
    try:
        elems = driver.find_elements(By.CSS_SELECTOR, "#content-holder-left h4 strong")
        for el in elems:
            if el.text.strip().lower() == label.lower():
                parent = el.find_element(By.XPATH, "./ancestor::h4")
                next_p = parent.find_element(By.XPATH, "following-sibling::p[1]")
                return next_p.text.strip()
    except Exception:
        return ""

def extract_description():
    """Extracts contract description."""
    try:
        paragraphs = driver.find_elements(By.XPATH, "//h3[text()='Description']/following-sibling::p")
        return "\n".join(p.text.strip() for p in paragraphs if p.text.strip())
    except:
        return ""

def extract_value(label):
    """Extracts numeric contract value safely."""
    text = find_value_by_label(label)
    if not text:
        return ""
    text = text.replace("£", "").replace(",", "")
    if "to" in text:
        return float(text.split("to")[0].strip())
    try:
        return float(text)
    except:
        return text

def get_website():
    """Extracts the Website field if available."""
    try:
        website_label = driver.find_element(By.XPATH, "//h4[strong[text()='Website']]")
        link = website_label.find_element(By.XPATH, "following-sibling::p[1]//a")
        return link.get_attribute("href").strip()
    except:
        return ""

# ===================== CORE SCRAPING LOGIC =====================

def fetch_data_from_link(link, retries=MAX_RETRIES):
    """Fetches contract and supplier data from a given link."""
    for attempt in range(retries):
        try:
            driver.get(link)
            WebDriverWait(driver, 10).until(  
                EC.presence_of_element_located((By.CSS_SELECTOR, "#all-content-wrapper"))
            )

            contract_data = [
                get_element_text("#all-content-wrapper > h1"),  # Title
                get_element_text("#home-breadcrumb-description > h2"),  # Buyer
                ", ".join([el.text.strip() for el in driver.find_elements(By.CSS_SELECTOR, "#content-holder-left > div:nth-child(3) > ul > li > p")]),  # Industry
                find_value_by_label("Location of contract"),
                extract_value("Total value of contract"),
                find_value_by_label("Procurement reference"),
                find_value_by_label("Published date"),
                find_value_by_label("Closing date"),
                find_value_by_label("Closing time"),
                find_value_by_label("Contract start date"),
                find_value_by_label("Contract end date"),
                find_value_by_label("Contract type"),
                find_value_by_label("Procedure type"),
                find_value_by_label("Contract is suitable for SMEs?"),
                find_value_by_label("Contract is suitable for VCSEs?"),
                extract_description(),
                find_value_by_label("Awarded date"),
                find_value_by_label("Contact name"),
                find_value_by_label("Address").replace("\n", ", "),
                get_element_text("a[href^='mailto:']"),
                get_website()
            ]

            # ✅ Extract First Supplier Using Fixed Selectors
            supplier = (
                get_element_text("#content-holder-left > div:nth-child(5) > h4:nth-child(12) > strong")
                or get_element_text("#content-holder-left > div:nth-child(6) > h4:nth-child(12) > strong")
            )

            supplier_address, reference, supplier_sme, supplier_vcse = "", "", "", ""

            # ✅ Click "Show Supplier Information" if available
            try:
                show_supplier_button = driver.find_element(By.CSS_SELECTOR, "#show_supplier_0_information_link > span")
                driver.execute_script("arguments[0].click();", show_supplier_button)
                time.sleep(1)

                supplier_address = get_element_text("#supplier_block_0 > dl > dd:nth-child(2) > p")
                reference = get_element_text("#supplier_block_0 > dl > dd:nth-child(4) > p")
                supplier_sme = get_element_text("#supplier_block_0 > dl > dd:nth-child(6) > p")
                supplier_vcse = get_element_text("#supplier_block_0 > dl > dd:nth-child(8) > p")
            except NoSuchElementException:
                pass

            contract_data.extend([supplier, supplier_address, reference, supplier_sme, supplier_vcse])

            return contract_data

        except (TimeoutException, NoSuchElementException, InvalidSessionIdException) as e:
            print(f"Attempt {attempt + 1} failed for {link}: {e}")
            if attempt == retries - 1:
                return None
            time.sleep(3)

# ===================== SAVE TO EXCEL =====================

def save_to_excel(record):
    try:
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb.active
        ws.append(record)
        wb.save(OUTPUT_EXCEL)
        print(f"✅ Saved: {record[0]}")
    except Exception as e:
        print(f"⚠️ Error saving to Excel: {e}")

# ===================== MAIN FUNCTION =====================

def main():
    df_links = pd.read_excel(INPUT_EXCEL)
    links = df_links["Links"].tolist()

    for i, link in enumerate(links):
        print(f"🔄 Processing {i+1}/{len(links)}: {link}")
        record = fetch_data_from_link(link)
        if record:
            save_to_excel(record)

    driver.quit()

if __name__ == "__main__":
    main()
